import re
import pandas as pd
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
import openpyxl
from datetime import datetime
import base64
import io
import os

DATA_FILE = os.path.join("data", "mandat sanitaire 2022.xlsx")

def get_file_mtime(path: str) -> float:
    try:
        return os.path.getmtime(path)
    except FileNotFoundError:
        return 0.0

@st.cache_data
def load_vaccination_data_from_path(path: str, mtime: float):
    # on réutilise ta fonction existante (load_vaccination_data) en lui passant le path
    return load_vaccination_data(path)

# ---------------------------
# CONFIGURATION
# ---------------------------
st.set_page_config(
    page_title="Mandat Sanitaire - Dashboard Vétérinaire",
    layout="wide",
    page_icon="🐾",
    initial_sidebar_state="collapsed"
)

# ---------------------------
# PRIX PAR DÉFAUT (modifiable par l'utilisateur)
# ---------------------------
PRIX_DEFAULT = {
    "aphto_ovin_caprin": {"prix_ovin": 0.0, "prix_caprin": 0.0},
    "ovin_clavelee": {"prix_ovin": 0.0},
    "bovin_aphto": {"prix_bovin": 0.0},
    "rage": {"prix_chien": 0.0},
}
# ---------------------------
# PALETTE BLEUE (logo vétérinaire)
# ---------------------------
BLUE_MAIN = "#1976d2"
BLUE_DARK = "#0d47a1"
BLUE_LIGHT = "#42a5f5"
BLUE_VERY_LIGHT = "#bbdefb"

CHART_COLORS = [BLUE_DARK, BLUE_MAIN, BLUE_LIGHT, "#90caf9", "#64b5f6"]
CHART_GRADIENT = [BLUE_VERY_LIGHT, BLUE_MAIN]

# ---------------------------
# FONCTIONS UTILITAIRES
# ---------------------------
def load_css(path="style.css"):
    with open(path, "r", encoding="utf-8") as f:
        st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)

def apply_transparent_theme(fig):
    fig.update_layout(
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)"
    )
    return fig

def kpi_cards(items):
    """Afficher des cartes KPI"""
    html = '<div class="kpi-grid">'
    for it in items:
        label = it.get("label", "")
        value = it.get("value", "")
        delta = it.get("delta", "")
        html += f"""
<div class="kpi-card">
<div class="kpi-label">{label}</div>
<div class="kpi-value">{value}</div>
<div class="kpi-delta">{delta}</div>
</div>
        """
    html += "</div>"
    st.markdown(html, unsafe_allow_html=True)

def apply_date_filter(df: pd.DataFrame, key: str) -> pd.DataFrame:
    """Filtre un dataframe par plage de dates Streamlit (si la colonne date existe)."""
    if df.empty or 'date' not in df.columns or df['date'].isna().all():
        return df

    # bornes min/max disponibles
    min_date = df['date'].min().date()
    max_date = df['date'].max().date()

    date_range = st.date_input(
        "Période de vaccination",
        value=(min_date, max_date),
        min_value=min_date,
        max_value=max_date,
        key=key
    )

    # Streamlit peut renvoyer une date unique ou un tuple (start, end)
    if isinstance(date_range, tuple) and len(date_range) == 2:
        start_date, end_date = date_range

        start_dt = pd.to_datetime(start_date)
        end_dt = pd.to_datetime(end_date) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)

        return df[(df['date'] >= start_dt) & (df['date'] <= end_dt)]

    return df

def reset_prix():
    # Reset des prix en session
    st.session_state.prix = {k: v.copy() for k, v in PRIX_DEFAULT.items()}
    st.session_state.prix_version = st.session_state.get("prix_version", 0) + 1

# ---------------------------
# CHARGEMENT DES DONNÉES
# ---------------------------
@st.cache_data
def load_vaccination_data(file_obj):
    """Charger les données de vaccination depuis le fichier Excel"""
    if isinstance(file_obj, str):
        wb = openpyxl.load_workbook(file_obj, data_only=True)
    elif isinstance(file_obj, (bytes, bytearray)):
        wb = openpyxl.load_workbook(io.BytesIO(file_obj), data_only=True)
    else:
        try:
            data = file_obj.getvalue()
        except Exception:
            data = file_obj.read()
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    
    datasets = {}
    
    # Charger aphto ovin et caprin
    ws = wb['aphto ovin et caprin']
    rows = list(ws.iter_rows(values_only=True))
    data = []
    for i, row in enumerate(rows):
        if i >= 2:  # Skip first 2 rows
            if row[11] and row[10]:  # Check for name and ID
                data.append({
                    'nom': row[11],
                    'cin': row[10],
                    'region': row[9],
                    'date': row[8],
                    'recu_num': row[7],
                    'ovins_vaccines': row[6] or 0,
                    'caprins_vaccines': row[5] or 0,
                    'total_ovins': row[4] or 0,
                    'total_caprins': row[3] or 0,
                })
    datasets['aphto_ovin_caprin'] = pd.DataFrame(data)
    
    # Charger ovin clavelee
    ws = wb['ovin clavelee']
    rows = list(ws.iter_rows(values_only=True))
    data = []
    for i, row in enumerate(rows):
        if i >= 3:  # Skip first 3 rows
            if row[8] and row[7]:  # Check for name and ID
                data.append({
                    'nom': row[8],
                    'cin': row[7],
                    'region': row[6],
                    'date': row[5],
                    'recu_num': row[4],
                    'ovins_vaccines': row[3] or 0,
                    'total_ovins': row[2] or 0,
                })
    datasets['ovin_clavelee'] = pd.DataFrame(data)
    
    # Charger bovin aphto
    ws = wb['bovin aphto']
    rows = list(ws.iter_rows(values_only=True))
    data = []
    for i, row in enumerate(rows):
        if i >= 3:  # Skip first 3 rows
            if row[10] and row[9]:  # Check for name and ID
                data.append({
                    'nom': row[10],
                    'cin': row[9],
                    'region': row[8],
                    'date': row[7],
                    'recu_num': row[6],
                    'bovins_vaccines': row[5] or 0,
                    'total_bovins': row[4] or 0,
                })
    datasets['bovin_aphto'] = pd.DataFrame(data)
    
    # Charger rage (داء الكلب)
    ws = wb['داء الكلب']
    rows = list(ws.iter_rows(values_only=True))
    data = []
    for i, row in enumerate(rows):
        if i >= 4:  # Skip first 4 rows
            if row[9] and row[8]:  # Check for name and ID
                data.append({
                    'nom': row[9],
                    'cin': row[8],
                    'region': row[7],
                    'date': row[6],
                    'recu_num': row[5],
                    'chiens_vaccines': row[4] or 0,
                    'total_chiens': row[3] or 0,
                })
    datasets['rage'] = pd.DataFrame(data)
    # --- NORMALISATION DES DATES (IMPORTANT POUR LES FILTRES) ---
    for k, df in datasets.items():
        if not df.empty and 'date' in df.columns:
            df['date'] = pd.to_datetime(df['date'], errors='coerce')
            datasets[k] = df

    return datasets

# ---------------------------
# CHARGEMENT CSS
# ---------------------------
load_css("style.css")

# ---------------------------
# PAGE D'UPLOAD OU DASHBOARD
# ---------------------------
if 'data_loaded' not in st.session_state:
    st.session_state.data_loaded = False
if "prix" not in st.session_state:
    st.session_state.prix = {k: v.copy() for k, v in PRIX_DEFAULT.items()}

# au lieu de if not data_loaded ... uploader ...
if not os.path.exists(DATA_FILE):
    st.error(f"Fichier introuvable: {DATA_FILE}")
    st.stop()

datasets = load_vaccination_data_from_path(DATA_FILE, get_file_mtime(DATA_FILE))
st.session_state.datasets = datasets
st.session_state.data_loaded = True

# Dashboard principal
# ---------------------------
# EN-TÊTE
# ---------------------------
st.markdown("""
<div class="vet-header">
    <div class="vet-logo-section">
        <div class="vet-icon">🐾</div>
        <div>
            <div class="vet-title">Dashboard Mandat Sanitaire 2022</div>
            <div class="vet-subtitle">Campagnes de Vaccination - Région de Sousse</div>
        </div>
    </div>
</div>
""", unsafe_allow_html=True)
datasets = st.session_state.datasets
# ---------------------------
# TABS PRINCIPALES
# ---------------------------
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "🐑 Fièvre Aphteuse (Ovins/Caprins)",
    "🐏 Clavelée des Ovins",
    "🐄 Fièvre Aphteuse (Bovins)",
    "🐕 Rage Canine",
    "🧮 Calculatrice",
])

# ---------------------------
# TAB 1: APHTO OVIN ET CAPRIN
# ---------------------------
with tab1:
    df = datasets['aphto_ovin_caprin']
    
    if len(df) > 0:
        # Filtres

        st.markdown("""
        <div class="section-head">
        <div class="section-icon-pro">🔍</div>
        <div class="section-title-pro">Filtres</div>
        <div class="section-line-pro"></div>
        </div>
        """, unsafe_allow_html=True)
        col1, col2 = st.columns(2)

        with col1:
            regions = sorted(df['region'].dropna().unique())
            selected_regions = st.multiselect("Région (العمادة)", regions, key="aphto_oc_region")

        # Application des filtres (région d'abord)
        filtered_df = df.copy()
        if selected_regions:
            filtered_df = filtered_df[filtered_df['region'].isin(selected_regions)]

        with col2:
            filtered_df = apply_date_filter(filtered_df, key="aphto_oc_dates")

    
        # KPIs
        total_ovins = int(filtered_df['total_ovins'].sum())
        total_caprins = int(filtered_df['total_caprins'].sum())
        ovins_vaccines = int(filtered_df['ovins_vaccines'].sum())
        caprins_vaccines = int(filtered_df['caprins_vaccines'].sum())
        total_animaux = total_ovins + total_caprins
        total_vaccines = ovins_vaccines + caprins_vaccines
        taux_vaccination = (total_vaccines / total_animaux * 100) if total_animaux > 0 else 0
        nb_eleveurs = len(filtered_df)
        
        kpi_cards([
            {"label": "Total Animaux", "value": f"{total_animaux:,}".replace(",", " "), "delta": "🐑 Ovins + Caprins"},
            {"label": "Animaux Vaccinés", "value": f"{total_vaccines:,}".replace(",", " "), "delta": f"✅ {taux_vaccination:.1f}% du total"},
            {"label": "Éleveurs Traités", "value": f"{nb_eleveurs:,}".replace(",", " "), "delta": "👨‍🌾 Bénéficiaires"},
            {"label": "Ovins", "value": f"{ovins_vaccines:,}/{total_ovins:,}".replace(",", " "), "delta": f"📊 {(ovins_vaccines/total_ovins*100 if total_ovins>0 else 0):.1f}%"},
            {"label": "Caprins", "value": f"{caprins_vaccines:,}/{total_caprins:,}".replace(",", " "), "delta": f"📊 {(caprins_vaccines/total_caprins*100 if total_caprins>0 else 0):.1f}%"},
        ])
        
        # Graphiques
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("""
                <div class="section-head">
                <div class="section-icon-pro">📊</div>
                <div class="section-title-pro">Répartition par région</div>
                <div class="section-line-pro"></div>
                </div>
                """, unsafe_allow_html=True)

            region_data = filtered_df.groupby('region').agg({
                'ovins_vaccines': 'sum',
                'caprins_vaccines': 'sum'
            }).reset_index()
            region_data['total'] = region_data['ovins_vaccines'] + region_data['caprins_vaccines']
            region_data = region_data.sort_values('total', ascending=False).head(10)
            
            fig = px.bar(
                region_data,
                x='total',
                y='region',
                orientation='h',
                color='total',
                color_continuous_scale=CHART_GRADIENT
            )
            fig.update_layout(height=400, margin=dict(l=10, r=10, t=10, b=10), showlegend=False)
            apply_transparent_theme(fig)
            st.plotly_chart(fig, use_container_width=True)
        
        with col2:
            st.markdown("""
                <div class="section-head">
                <div class="section-icon-pro">🐑</div>
                <div class="section-title-pro"> Ovins vs Caprins vaccinés</div>
                <div class="section-line-pro"></div>
                </div>
                """, unsafe_allow_html=True)            
            species_data = pd.DataFrame({
                'Type': ['Ovins', 'Caprins'],
                'Vaccinés': [ovins_vaccines, caprins_vaccines]
            })
            
            fig = px.pie(
                species_data,
                values='Vaccinés',
                names='Type',
                color_discrete_sequence=CHART_COLORS[:2]
            )
            fig.update_layout(height=400, margin=dict(l=10, r=10, t=10, b=10))
            fig.update_traces(textposition='inside', textinfo='percent+label+value')
            apply_transparent_theme(fig)
            st.plotly_chart(fig, use_container_width=True)
        
        # Evolution temporelle
        if 'date' in filtered_df.columns and filtered_df['date'].notna().any():
            st.markdown("""
                <div class="section-head">
                <div class="section-icon-pro">📈</div>
                <div class="section-title-pro">Évolution temporelle des vaccinations</div>
                <div class="section-line-pro"></div>
                </div>
                """, unsafe_allow_html=True)          
            temporal_data = filtered_df.copy()
            temporal_data['date'] = pd.to_datetime(temporal_data['date'])
            temporal_data = temporal_data.groupby(temporal_data['date'].dt.date).agg({
                'ovins_vaccines': 'sum',
                'caprins_vaccines': 'sum'
            }).reset_index()
            temporal_data['total'] = temporal_data['ovins_vaccines'] + temporal_data['caprins_vaccines']
            temporal_data = temporal_data.sort_values('date')
            
            fig = go.Figure()
            fig.add_trace(go.Scatter(
                x=temporal_data['date'],
                y=temporal_data['total'],
                mode='lines+markers',
                name='Total',
                line=dict(color=BLUE_MAIN, width=3),                    
                marker=dict(size=8, color=BLUE_DARK)
            ))
            fig.update_layout(height=350, margin=dict(l=10, r=10, t=10, b=10))
            apply_transparent_theme(fig)
            st.plotly_chart(fig, use_container_width=True)
        
        # Tableau détaillé
        st.markdown("""
                <div class="section-head">
                <div class="section-icon-pro">📋</div>
                <div class="section-title-pro">Liste des éleveurs</div>
                <div class="section-line-pro"></div>
                </div>
                """, unsafe_allow_html=True)   
        display_cols = ['nom', 'region', 'date', 'ovins_vaccines', 'total_ovins', 'caprins_vaccines', 'total_caprins']
        st.dataframe(filtered_df[display_cols], use_container_width=True, height=400)
    else:
        st.warning("Aucune donnée disponible pour cette campagne.")

# ---------------------------
# TAB 2: OVIN CLAVELEE
# ---------------------------
with tab2:
    df = datasets['ovin_clavelee']
    
    if len(df) > 0:
        # Filtres
        st.markdown("""
        <div class="section-head">
        <div class="section-icon-pro">🔍</div>
        <div class="section-title-pro">Filtres</div>
        <div class="section-line-pro"></div>
        </div>
        """, unsafe_allow_html=True)
        col1, col2 = st.columns(2)

        with col1:
            regions = sorted(df['region'].dropna().unique())
            selected_regions = st.multiselect("Région (العمادة)", regions, key="clavelee_region")

        # Application filtre région
        filtered_df = df.copy()
        if selected_regions:
            filtered_df = filtered_df[filtered_df['region'].isin(selected_regions)]

        with col2:
            filtered_df = apply_date_filter(filtered_df, key="clavelee_dates")

        
        # KPIs
        total_ovins = int(filtered_df['total_ovins'].sum())
        ovins_vaccines = int(filtered_df['ovins_vaccines'].sum())
        taux_vaccination = (ovins_vaccines / total_ovins * 100) if total_ovins > 0 else 0
        nb_eleveurs = len(filtered_df)
        
        kpi_cards([
            {"label": "Total Ovins", "value": f"{total_ovins:,}".replace(",", " "), "delta": "🐏 Population totale"},
            {"label": "Ovins Vaccinés", "value": f"{ovins_vaccines:,}".replace(",", " "), "delta": f"✅ {taux_vaccination:.1f}% du total"},
            {"label": "Éleveurs Traités", "value": f"{nb_eleveurs:,}".replace(",", " "), "delta": "👨‍🌾 Bénéficiaires"},
            {"label": "Taux de Vaccination", "value": f"{taux_vaccination:.1f}%", "delta": "📈 Couverture"},
        ])
        
        # Graphiques
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("""
            <div class="section-head">
            <div class="section-icon-pro">📊</div>
            <div class="section-title-pro">Répartition par région</div>
            <div class="section-line-pro"></div>
            </div>
            """, unsafe_allow_html=True)
            region_data = filtered_df.groupby('region')['ovins_vaccines'].sum().reset_index()
            region_data = region_data.sort_values('ovins_vaccines', ascending=False).head(10)
            
            fig = px.bar(
                region_data,
                x='ovins_vaccines',
                y='region',
                orientation='h',
                color='ovins_vaccines',
                color_continuous_scale=CHART_GRADIENT
            )
            fig.update_layout(height=400, margin=dict(l=10, r=10, t=10, b=10), showlegend=False)
            apply_transparent_theme(fig)
            st.plotly_chart(fig, use_container_width=True)
        
        with col2:
            st.markdown("""
            <div class="section-head">
            <div class="section-icon-pro">📊</div>
            <div class="section-title-pro">Taux de vaccination par région</div>
            <div class="section-line-pro"></div>
            </div>
            """, unsafe_allow_html=True)

            region_data = filtered_df.groupby('region').agg({
                'ovins_vaccines': 'sum',
                'total_ovins': 'sum'
            }).reset_index()
            region_data['taux'] = (region_data['ovins_vaccines'] / region_data['total_ovins'] * 100).round(1)
            region_data = region_data.sort_values('taux', ascending=False).head(10)
            
            fig = px.bar(
                region_data,
                x='taux',
                y='region',
                orientation='h',
                color='taux',
                color_continuous_scale=CHART_GRADIENT
            )
            fig.update_layout(height=400, margin=dict(l=10, r=10, t=10, b=10), showlegend=False)
            apply_transparent_theme(fig)
            st.plotly_chart(fig, use_container_width=True)
        
        # Tableau détaillé
        st.markdown("""
                <div class="section-head">
                <div class="section-icon-pro">📋</div>
                <div class="section-title-pro">Liste des éleveurs</div>
                <div class="section-line-pro"></div>
                </div>
                """, unsafe_allow_html=True)   
        display_cols = ['nom', 'region', 'date', 'ovins_vaccines', 'total_ovins']
        st.dataframe(filtered_df[display_cols], use_container_width=True, height=400)
    else:
        st.warning("Aucune donnée disponible pour cette campagne.")

# ---------------------------
# TAB 3: BOVIN APHTO
# ---------------------------
with tab3:
    df = datasets['bovin_aphto']
    
    if len(df) > 0:
        # Filtres
        st.markdown("""
        <div class="section-head">
        <div class="section-icon-pro">🔍</div>
        <div class="section-title-pro">Filtres</div>
        <div class="section-line-pro"></div>
        </div>
        """, unsafe_allow_html=True)

        col1, col2 = st.columns(2)

        with col1:
            regions = sorted(df['region'].dropna().unique())
            selected_regions = st.multiselect("Région (العمادة)", regions, key="bovin_region")

        # Application filtre région
        filtered_df = df.copy()
        if selected_regions:
            filtered_df = filtered_df[filtered_df['region'].isin(selected_regions)]

        with col2:
            filtered_df = apply_date_filter(filtered_df, key="bovin_dates")

        # KPIs
        total_bovins = int(filtered_df['total_bovins'].sum())
        bovins_vaccines = int(filtered_df['bovins_vaccines'].sum())
        taux_vaccination = (bovins_vaccines / total_bovins * 100) if total_bovins > 0 else 0
        nb_eleveurs = len(filtered_df)
        
        kpi_cards([
            {"label": "Total Bovins", "value": f"{total_bovins:,}".replace(",", " "), "delta": "🐄 Population totale"},
            {"label": "Bovins Vaccinés", "value": f"{bovins_vaccines:,}".replace(",", " "), "delta": f"✅ {taux_vaccination:.1f}% du total"},
            {"label": "Éleveurs Traités", "value": f"{nb_eleveurs:,}".replace(",", " "), "delta": "👨‍🌾 Bénéficiaires"},
            {"label": "Taux de Vaccination", "value": f"{taux_vaccination:.1f}%", "delta": "📈 Couverture"},
        ])
        
        # Graphiques
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("""
            <div class="section-head">
            <div class="section-icon-pro">📊</div>
            <div class="section-title-pro">Répartition par région</div>
            <div class="section-line-pro"></div>
            </div>
            """, unsafe_allow_html=True)
            region_data = filtered_df.groupby('region')['bovins_vaccines'].sum().reset_index()
            region_data = region_data.sort_values('bovins_vaccines', ascending=False).head(10)
            
            fig = px.bar(
                region_data,
                x='bovins_vaccines',
                y='region',
                orientation='h',
                color='bovins_vaccines',
                color_continuous_scale=CHART_GRADIENT
            )
            fig.update_layout(height=400, margin=dict(l=10, r=10, t=10, b=10), showlegend=False)
            apply_transparent_theme(fig)
            st.plotly_chart(fig, use_container_width=True)
        
        with col2:
            st.markdown("""
            <div class="section-head">
            <div class="section-icon-pro">📊</div>
            <div class="section-title-pro">Distribution des tailles de troupeaux</div>
            <div class="section-line-pro"></div>
            </div>
            """, unsafe_allow_html=True)
            fig = px.histogram(
                filtered_df,
                x='total_bovins',
                nbins=20,
                color_discrete_sequence=[BLUE_MAIN]
            )
            fig.update_layout(height=400, margin=dict(l=10, r=10, t=10, b=10))
            apply_transparent_theme(fig)
            st.plotly_chart(fig, use_container_width=True)
        
        # Tableau détaillé
        st.markdown("""
                <div class="section-head">
                <div class="section-icon-pro">📋</div>
                <div class="section-title-pro">Liste des éleveurs</div>
                <div class="section-line-pro"></div>
                </div>
                """, unsafe_allow_html=True)   
        display_cols = ['nom', 'region', 'date', 'bovins_vaccines', 'total_bovins']
        st.dataframe(filtered_df[display_cols], use_container_width=True, height=400)
    else:
        st.warning("Aucune donnée disponible pour cette campagne.")

# ---------------------------
# TAB 4: RAGE
# ---------------------------
with tab4:
    df = datasets['rage']
    
    if len(df) > 0:
        # Filtres
        st.markdown("""
        <div class="section-head">
        <div class="section-icon-pro">🔍</div>
        <div class="section-title-pro">Filtres</div>
        <div class="section-line-pro"></div>
        </div>
        """, unsafe_allow_html=True)
        col1, col2 = st.columns(2)

        with col1:
            regions = sorted(df['region'].dropna().unique())
            selected_regions = st.multiselect("Région (العمادة)", regions, key="rage_region")

        # Application filtre région
        filtered_df = df.copy()
        if selected_regions:
            filtered_df = filtered_df[filtered_df['region'].isin(selected_regions)]

        with col2:
            filtered_df = apply_date_filter(filtered_df, key="rage_dates")

        # KPIs
        total_chiens = int(filtered_df['total_chiens'].sum())
        chiens_vaccines = int(filtered_df['chiens_vaccines'].sum())
        taux_vaccination = (chiens_vaccines / total_chiens * 100) if total_chiens > 0 else 0
        nb_proprietaires = len(filtered_df)
        
        kpi_cards([
            {"label": "Total Chiens", "value": f"{total_chiens:,}".replace(",", " "), "delta": "🐕 Population totale"},
            {"label": "Chiens Vaccinés", "value": f"{chiens_vaccines:,}".replace(",", " "), "delta": f"✅ {taux_vaccination:.1f}% du total"},
            {"label": "Propriétaires", "value": f"{nb_proprietaires:,}".replace(",", " "), "delta": "👤 Bénéficiaires"},
            {"label": "Taux de Vaccination", "value": f"{taux_vaccination:.1f}%", "delta": "📈 Couverture"},
        ])
        
        # Graphiques
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("""
            <div class="section-head">
            <div class="section-icon-pro">📊</div>
            <div class="section-title-pro">Répartition par région</div>
            <div class="section-line-pro"></div>
            </div>
            """, unsafe_allow_html=True)
            region_data = filtered_df.groupby('region')['chiens_vaccines'].sum().reset_index()
            region_data = region_data.sort_values('chiens_vaccines', ascending=False).head(10)
            
            fig = px.bar(
                region_data,
                x='chiens_vaccines',
                y='region',
                orientation='h',
                color='chiens_vaccines',
                color_continuous_scale=CHART_GRADIENT
            )
            fig.update_layout(height=400, margin=dict(l=10, r=10, t=10, b=10), showlegend=False)
            apply_transparent_theme(fig)
            st.plotly_chart(fig, use_container_width=True)
        
        with col2:
            st.markdown("""
            <div class="section-head">
            <div class="section-icon-pro">🏘️</div>
            <div class="section-title-pro">Top 10 régions</div>
            <div class="section-line-pro"></div>
            </div>
            """, unsafe_allow_html=True)
            region_data = filtered_df.groupby('region').agg({
                'chiens_vaccines': 'sum',
                'total_chiens': 'sum'
            }).reset_index()
            region_data['taux'] = (region_data['chiens_vaccines'] / region_data['total_chiens'] * 100).round(1)
            region_data = region_data.sort_values('chiens_vaccines', ascending=False).head(10)
            
            fig = px.pie(
                region_data,
                values='chiens_vaccines',
                names='region',
                color_discrete_sequence=px.colors.sequential.Blues
            )
            fig.update_layout(height=400, margin=dict(l=10, r=10, t=10, b=10))
            fig.update_traces(textposition='inside', textinfo='percent+label')
            apply_transparent_theme(fig)
            st.plotly_chart(fig, use_container_width=True)
        
        # Tableau détaillé
        st.markdown("""
                <div class="section-head">
                <div class="section-icon-pro">📋</div>
                <div class="section-title-pro">Liste des éleveurs</div>
                <div class="section-line-pro"></div>
                </div>
                """, unsafe_allow_html=True)   
        display_cols = ['nom', 'region', 'date', 'chiens_vaccines', 'total_chiens']
        st.dataframe(filtered_df[display_cols], use_container_width=True, height=400)
    else:
        st.warning("Aucune donnée disponible pour cette campagne.")
# ---------------------------
# TAB 5: CALCULATRICE
# ---------------------------
with tab5:
    # Configuration des campagnes
    type_options = {
        "aphto_ovin_caprin": {"label": "Fièvre Aphteuse (Ovins/Caprins)", "icon": "🐑🐐", "color": "#1976d2"},
        "ovin_clavelee": {"label": "Clavelée des Ovins", "icon": "🐏", "color": "#0d47a1"},
        "bovin_aphto": {"label": "Fièvre Aphteuse (Bovins)", "icon": "🐄", "color": "#2196f3"},
        "rage": {"label": "Rage Canine", "icon": "🐕", "color": "#1565c0"},
    }

    # Init version
    if "prix_version" not in st.session_state:
        st.session_state.prix_version = 0
    v = st.session_state.prix_version

    # Sélection du type
    selected_key = st.selectbox(
        "Type de vaccination",
        options=list(type_options.keys()),
        format_func=lambda k: type_options[k]["label"],
        key="calc_type",
        label_visibility="collapsed"
    )

    # Variables pour le template
    campaign_label = type_options[selected_key]["label"]
    campaign_icon = type_options[selected_key]["icon"]
    campaign_color = type_options[selected_key]["color"]

    # Section Paramètres des prix
    st.markdown(f"""
    <div style="margin: 2rem 0 1.5rem 0;">
        <div style="display: flex; align-items: center; gap: 1rem; margin-bottom: 1.5rem;">
            <div style="font-size: 32px; background: linear-gradient(135deg, {campaign_color}, #42a5f5); 
                 width: 56px; height: 56px; border-radius: 14px; display: flex; align-items: center; 
                 justify-content: center; box-shadow: 0 4px 16px rgba(25, 118, 210, 0.25);">💵</div>
            <div style="color: #0d47a1; font-size: 22px; font-weight: 800;">Paramètres des Prix</div>
            <div style="flex: 1; height: 3px; background: linear-gradient(90deg, {campaign_color} 0%, transparent 100%); border-radius: 2px;"></div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Prix actuels en session
    if "prix" not in st.session_state:
        st.session_state.prix = {k: v.copy() for k, v in PRIX_DEFAULT.items()}

    prix = st.session_state.prix.get(selected_key, {}).copy()

    p1, p2, p3 = st.columns([1, 1, 0.8])

    # Widgets de prix selon la campagne
    if selected_key == "aphto_ovin_caprin":
        with p1:
            prix_ovin = st.number_input(
                "💰 Prix par ovin vacciné",
                min_value=0.0,
                value=float(prix.get("prix_ovin", 0.0)),
                step=0.1,
                key=f"ui_prix_ovin_aphto_oc__v{v}"
            )
        with p2:
            prix_caprin = st.number_input(
                "💰 Prix par caprin vacciné",
                min_value=0.0,
                value=float(prix.get("prix_caprin", 0.0)),
                step=0.1,
                key=f"ui_prix_caprin_aphto_oc__v{v}"
            )
        prix["prix_ovin"] = float(prix_ovin)
        prix["prix_caprin"] = float(prix_caprin)
        
        with p3:
            st.markdown("<div style='margin-top:25px;'></div>", unsafe_allow_html=True)
            if st.button("🔄 Réinitialiser", key="btn_reset_prix_unique"):
                reset_prix()
                st.rerun()

    elif selected_key == "ovin_clavelee":
        with p1:
            prix_ovin = st.number_input(
                "💰 Prix par ovin vacciné",
                min_value=0.0,
                value=float(prix.get("prix_ovin", 0.0)),
                step=0.1,
                key=f"ui_prix_ovin_clavelee__v{v}"
            )
        prix["prix_ovin"] = float(prix_ovin)
        
        with p2:
            st.markdown("<div style='margin-top:25px;'></div>", unsafe_allow_html=True)
            if st.button("🔄 Réinitialiser", key="btn_reset_prix_unique"):
                reset_prix()
                st.rerun()
                
    elif selected_key == "bovin_aphto":
        with p1:
            prix_bovin = st.number_input(
                "💰 Prix par bovin vacciné",
                min_value=0.0,
                value=float(prix.get("prix_bovin", 0.0)),
                step=0.1,
                key=f"ui_prix_bovin__v{v}"
            )
        prix["prix_bovin"] = float(prix_bovin)
        
        with p2:
            st.markdown("<div style='margin-top:25px;'></div>", unsafe_allow_html=True)
            if st.button("🔄 Réinitialiser", key="btn_reset_prix_unique"):
                reset_prix()
                st.rerun()
                
    else:  # rage
        with p1:
            prix_chien = st.number_input(
                "💰 Prix par chien vacciné",
                min_value=0.0,
                value=float(prix.get("prix_chien", 0.0)),
                step=0.1,
                key=f"ui_prix_chien__v{v}"
            )
        prix["prix_chien"] = float(prix_chien)
        
        with p2:
            st.markdown("<div style='margin-top:25px;'></div>", unsafe_allow_html=True)
            if st.button("🔄 Réinitialiser", key="btn_reset_prix_unique"):
                reset_prix()
                st.rerun()

    # Sauvegarder prix en session
    st.session_state.prix[selected_key] = prix

    # Section Filtres
    df = datasets[selected_key]
    if df.empty:
        st.markdown("""
        <div style="background: linear-gradient(135deg, #fef3c7 0%, #fef9e7 100%); 
             border: 2px solid #fbbf24; border-radius: 16px; padding: 1.5rem; 
             margin: 2rem 0; text-align: center;">
            <div style="font-size: 48px; margin-bottom: 1rem;">⚠️</div>
            <div style="color: #92400e; font-size: 18px; font-weight: 700;">
                Aucune donnée disponible pour ce type de vaccination
            </div>
        </div>
        """, unsafe_allow_html=True)
    else:
        # Section Filtres
        st.markdown(f"""
        <div style="margin: 2.5rem 0 1.5rem 0;">
            <div style="display: flex; align-items: center; gap: 1rem; margin-bottom: 1.5rem;">
                <div style="font-size: 32px; background: linear-gradient(135deg, {campaign_color}, #42a5f5); 
                     width: 56px; height: 56px; border-radius: 14px; display: flex; align-items: center; 
                     justify-content: center; box-shadow: 0 4px 16px rgba(25, 118, 210, 0.25);">🔍</div>
                <div style="color: #0d47a1; font-size: 22px; font-weight: 800;">Filtres de Données</div>
                <div style="flex: 1; height: 3px; background: linear-gradient(90deg, {campaign_color} 0%, transparent 100%); border-radius: 2px;"></div>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        c1, c2 = st.columns(2)

        with c1:
            regions = sorted(df['region'].dropna().unique())
            selected_regions = st.multiselect("📍 Région (العمادة)", regions, key="calc_region")

        filtered_df = df.copy()
        if selected_regions:
            filtered_df = filtered_df[filtered_df["region"].isin(selected_regions)]

        with c2:
            filtered_df = apply_date_filter(filtered_df, key="calc_dates")

        # Calculs
        montant_total = 0.0
        details = []

        if selected_key == "aphto_ovin_caprin":
            prix_ovin = float(st.session_state.prix[selected_key]["prix_ovin"])
            prix_caprin = float(st.session_state.prix[selected_key]["prix_caprin"])
            nb_ovins = float(filtered_df["ovins_vaccines"].sum())
            nb_caprins = float(filtered_df["caprins_vaccines"].sum())
            montant_ovins = nb_ovins * prix_ovin
            montant_caprins = nb_caprins * prix_caprin
            montant_total = montant_ovins + montant_caprins
            details = [
                {"Espèce": "Ovins vaccinés 🐑", "Quantité": f"{nb_ovins:,.0f}".replace(",", " "), "Prix unitaire": f"{prix_ovin:.2f} DH", "Montant": f"{montant_ovins:,.2f} DH".replace(",", " ")},
                {"Espèce": "Caprins vaccinés 🐐", "Quantité": f"{nb_caprins:,.0f}".replace(",", " "), "Prix unitaire": f"{prix_caprin:.2f} DH", "Montant": f"{montant_caprins:,.2f} DH".replace(",", " ")},
            ]

        elif selected_key == "ovin_clavelee":
            prix_ovin = float(st.session_state.prix[selected_key]["prix_ovin"])
            nb_ovins = float(filtered_df["ovins_vaccines"].sum())
            montant_total = nb_ovins * prix_ovin
            details = [
                {"Espèce": "Ovins vaccinés 🐏", "Quantité": f"{nb_ovins:,.0f}".replace(",", " "), "Prix unitaire": f"{prix_ovin:.2f} DH", "Montant": f"{montant_total:,.2f} DH".replace(",", " ")},
            ]

        elif selected_key == "bovin_aphto":
            prix_bovin = float(st.session_state.prix[selected_key]["prix_bovin"])
            nb_bovins = float(filtered_df["bovins_vaccines"].sum())
            montant_total = nb_bovins * prix_bovin
            details = [
                {"Espèce": "Bovins vaccinés 🐄", "Quantité": f"{nb_bovins:,.0f}".replace(",", " "), "Prix unitaire": f"{prix_bovin:.2f} DH", "Montant": f"{montant_total:,.2f} DH".replace(",", " ")},
            ]

        else:  # rage
            prix_chien = float(st.session_state.prix[selected_key]["prix_chien"])
            nb_chiens = float(filtered_df["chiens_vaccines"].sum())
            montant_total = nb_chiens * prix_chien
            details = [
                {"Espèce": "Chiens vaccinés 🐕", "Quantité": f"{nb_chiens:,.0f}".replace(",", " "), "Prix unitaire": f"{prix_chien:.2f} DH", "Montant": f"{montant_total:,.2f} DH".replace(",", " ")},
            ]

        # Affichage KPI Cards
        kpi_cards([
            {"label": "Lignes filtrées", "value": f"{len(filtered_df):,}".replace(",", " "), "delta": "📌 Après filtres"},
            {"label": "Montant total", "value": f"{montant_total:,.2f} DH".replace(",", " "), "delta": "💰 Total à payer"},
        ])

        # Détail du calcul
        st.markdown(f"""
        <div style="margin: 2.5rem 0 1rem 0;">
            <div style="display: flex; align-items: center; gap: 1rem; margin-bottom: 1rem;">
                <div style="font-size: 28px; background: linear-gradient(135deg, {campaign_color}, #42a5f5); 
                     width: 48px; height: 48px; border-radius: 12px; display: flex; align-items: center; 
                     justify-content: center; box-shadow: 0 4px 16px rgba(25, 118, 210, 0.25);">📋</div>
                <div style="color: #0d47a1; font-size: 20px; font-weight: 800;">Détail du Calcul</div>
                <div style="flex: 1; height: 2px; background: linear-gradient(90deg, {campaign_color} 0%, transparent 100%);"></div>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        st.dataframe(pd.DataFrame(details), use_container_width=True, height=220)

        # Montant par région
        st.markdown(f"""
        <div style="margin: 2.5rem 0 1rem 0;">
            <div style="display: flex; align-items: center; gap: 1rem; margin-bottom: 1rem;">
                <div style="font-size: 28px; background: linear-gradient(135deg, {campaign_color}, #42a5f5); 
                     width: 48px; height: 48px; border-radius: 12px; display: flex; align-items: center; 
                     justify-content: center; box-shadow: 0 4px 16px rgba(25, 118, 210, 0.25);">🏷️</div>
                <div style="color: #0d47a1; font-size: 20px; font-weight: 800;">Montant par Région</div>
                <div style="flex: 1; height: 2px; background: linear-gradient(90deg, {campaign_color} 0%, transparent 100%);"></div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        if filtered_df.empty:
            st.markdown("""
            <div style="background: linear-gradient(135deg, #dbeafe 0%, #eff6ff 100%); 
                 border: 2px solid #60a5fa; border-radius: 16px; padding: 1.5rem; 
                 margin: 1rem 0; text-align: center;">
                <div style="font-size: 48px; margin-bottom: 1rem;">ℹ️</div>
                <div style="color: #1e40af; font-size: 16px; font-weight: 600;">
                    Aucune donnée disponible après application des filtres
                </div>
            </div>
            """, unsafe_allow_html=True)
        else:
            if selected_key == "aphto_ovin_caprin":
                prix_ovin = float(st.session_state.prix[selected_key]["prix_ovin"])
                prix_caprin = float(st.session_state.prix[selected_key]["prix_caprin"])
                by_region = filtered_df.groupby("region").agg(
                    ovins_vaccines=("ovins_vaccines", "sum"),
                    caprins_vaccines=("caprins_vaccines", "sum")
                ).reset_index()
                by_region["montant"] = by_region["ovins_vaccines"] * prix_ovin + by_region["caprins_vaccines"] * prix_caprin

            elif selected_key == "ovin_clavelee":
                prix_ovin = float(st.session_state.prix[selected_key]["prix_ovin"])
                by_region = filtered_df.groupby("region").agg(
                    ovins_vaccines=("ovins_vaccines", "sum")
                ).reset_index()
                by_region["montant"] = by_region["ovins_vaccines"] * prix_ovin

            elif selected_key == "bovin_aphto":
                prix_bovin = float(st.session_state.prix[selected_key]["prix_bovin"])
                by_region = filtered_df.groupby("region").agg(
                    bovins_vaccines=("bovins_vaccines", "sum")
                ).reset_index()
                by_region["montant"] = by_region["bovins_vaccines"] * prix_bovin

            else:  # rage
                prix_chien = float(st.session_state.prix[selected_key]["prix_chien"])
                by_region = filtered_df.groupby("region").agg(
                    chiens_vaccines=("chiens_vaccines", "sum")
                ).reset_index()
                by_region["montant"] = by_region["chiens_vaccines"] * prix_chien

            by_region = by_region.sort_values("montant", ascending=False)
            st.dataframe(by_region, use_container_width=True, height=300)

