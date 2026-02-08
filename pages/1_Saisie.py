import re
import pandas as pd
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
import openpyxl
from datetime import datetime
import base64
import io
import os
from copy import copy
from datetime import date, datetime

DATA_FILE = os.path.join("data", "mandat sanitaire 2026.xlsx")

# ---------------------------
# CHARGEMENT CSS
# ---------------------------
def load_css(path="style.css"):
    with open(path, "r", encoding="utf-8") as f:
        st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)

# Afficher un message après rerun
if st.session_state.get("save_ok", False):
    st.success(st.session_state.get("save_msg", "✅ Données enregistrées avec succès."))
    st.session_state["save_ok"] = False

load_css("style.css")

def _next_seq(ws, seq_col: int, start_row: int = 1) -> int:
    """Trouver le prochain numéro de séquence dans la colonne seq_col."""
    last = 0
    for r in range(ws.max_row, start_row - 1, -1):
        v = ws.cell(r, seq_col).value
        if v is not None and str(v).strip() != "":
            try:
                last = int(v)
            except Exception:
                last = 0
            break
    return last + 1

# ---------------------------
# Helpers
# ---------------------------
def find_last_data_row(ws, key_col: int, start_row: int) -> int:
    """Dernière ligne qui contient une valeur dans key_col (ignore les lignes juste formatées)."""
    r = ws.max_row
    while r >= start_row:
        v = ws.cell(r, key_col).value
        if v is not None and str(v).strip() != "":
            return r
        r -= 1
    return start_row - 1

def copy_row_style(ws, src_row: int, dst_row: int, max_col: int):
    """Copie le style (bordures/cadres, formats, etc.) de src_row vers dst_row."""
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for c in range(1, max_col + 1):
        src = ws.cell(src_row, c)
        dst = ws.cell(dst_row, c)
        if src.has_style:
            dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)

# ---------------------------
# NOUVELLE FONCTION: Lire les enregistrements
# ---------------------------
@st.cache_data
def load_records_from_excel(path: str, campaign: str):
    """Charge tous les enregistrements d'une campagne donnée"""
    wb = openpyxl.load_workbook(path, data_only=True)
    records = []
    
    if campaign == "aphto_ovin_caprin":
        ws = wb["aphto ovin et caprin"]
        start_row = 3
        
        for row_idx in range(start_row, ws.max_row + 1):
            seq = ws.cell(row_idx, 13).value  # M
            if seq is None or str(seq).strip() == "":
                continue
                
            rec = {
                "row_idx": row_idx,
                "seq": int(seq),
                "nom": ws.cell(row_idx, 12).value or "",  # L
                "cin": ws.cell(row_idx, 11).value or "",  # K
                "region": ws.cell(row_idx, 10).value or "",  # J
                "recu_num": ws.cell(row_idx, 8).value or "",  # H
                "date": ws.cell(row_idx, 9).value,  # I
                "total_caprins": ws.cell(row_idx, 4).value or 0,  # D
                "total_ovins": ws.cell(row_idx, 5).value or 0,  # E
                "caprins_vaccines": ws.cell(row_idx, 6).value or 0,  # F
                "ovins_vaccines": ws.cell(row_idx, 7).value or 0,  # G
            }
            records.append(rec)
            
    elif campaign == "ovin_clavelee":
        ws = wb["ovin clavelee"]
        start_row = 4
        
        for row_idx in range(start_row, ws.max_row + 1):
            seq = ws.cell(row_idx, 10).value  # J
            if seq is None or str(seq).strip() == "":
                continue
                
            rec = {
                "row_idx": row_idx,
                "seq": int(seq),
                "nom": ws.cell(row_idx, 9).value or "",  # I
                "cin": ws.cell(row_idx, 8).value or "",  # H
                "region": ws.cell(row_idx, 7).value or "",  # G
                "recu_num": ws.cell(row_idx, 5).value or "",  # E
                "date": ws.cell(row_idx, 6).value,  # F
                "total_ovins": ws.cell(row_idx, 3).value or 0,  # C
                "ovins_vaccines": ws.cell(row_idx, 4).value or 0,  # D
            }
            records.append(rec)
            
    elif campaign == "bovin_aphto":
        ws = wb["bovin aphto"]
        start_row = 4
        
        for row_idx in range(start_row, ws.max_row + 1):
            seq = ws.cell(row_idx, 12).value  # L
            if seq is None or str(seq).strip() == "":
                continue
                
            rec = {
                "row_idx": row_idx,
                "seq": int(seq),
                "nom": ws.cell(row_idx, 11).value or "",  # K
                "cin": ws.cell(row_idx, 10).value or "",  # J
                "region": ws.cell(row_idx, 9).value or "",  # I
                "recu_num": ws.cell(row_idx, 7).value or "",  # G
                "date": ws.cell(row_idx, 8).value,  # H
                "total_bovins": ws.cell(row_idx, 5).value or 0,  # E
                "bovins_vaccines": ws.cell(row_idx, 6).value or 0,  # F
            }
            records.append(rec)
            
    else:  # rage
        ws = wb["داء الكلب"]
        start_row = 5
        
        for row_idx in range(start_row, ws.max_row + 1):
            seq = ws.cell(row_idx, 11).value  # K
            if seq is None or str(seq).strip() == "":
                continue
                
            rec = {
                "row_idx": row_idx,
                "seq": int(seq),
                "nom": ws.cell(row_idx, 10).value or "",  # J
                "cin": ws.cell(row_idx, 9).value or "",  # I
                "region": ws.cell(row_idx, 8).value or "",  # H
                "recu_num": ws.cell(row_idx, 6).value or "",  # F
                "date": ws.cell(row_idx, 7).value,  # G
                "total_chiens": ws.cell(row_idx, 4).value or 0,  # D
                "chiens_vaccines": ws.cell(row_idx, 5).value or 0,  # E
            }
            records.append(rec)
    
    wb.close()
    return records

# ---------------------------
# NOUVELLE FONCTION: Modifier un enregistrement
# ---------------------------
def update_record_in_excel(path: str, campaign: str, row_idx: int, rec: dict):
    """Modifie un enregistrement existant dans Excel"""
    wb = openpyxl.load_workbook(path)
    
    if campaign == "aphto_ovin_caprin":
        ws = wb["aphto ovin et caprin"]
        ws.cell(row_idx, 4).value = int(rec["total_caprins"])
        ws.cell(row_idx, 5).value = int(rec["total_ovins"])
        ws.cell(row_idx, 6).value = int(rec["caprins_vaccines"])
        ws.cell(row_idx, 7).value = int(rec["ovins_vaccines"])
        ws.cell(row_idx, 8).value = rec["recu_num"]
        ws.cell(row_idx, 9).value = rec["date"]
        ws.cell(row_idx, 9).number_format = "DD/MM/YYYY"
        ws.cell(row_idx, 10).value = rec["region"]
        ws.cell(row_idx, 11).value = rec["cin"]
        ws.cell(row_idx, 12).value = rec["nom"]
        
    elif campaign == "ovin_clavelee":
        ws = wb["ovin clavelee"]
        ws.cell(row_idx, 3).value = int(rec["total_ovins"])
        ws.cell(row_idx, 4).value = int(rec["ovins_vaccines"])
        ws.cell(row_idx, 5).value = rec["recu_num"]
        ws.cell(row_idx, 6).value = rec["date"]
        ws.cell(row_idx, 6).number_format = "DD/MM/YYYY"
        ws.cell(row_idx, 7).value = rec["region"]
        ws.cell(row_idx, 8).value = rec["cin"]
        ws.cell(row_idx, 9).value = rec["nom"]
        
    elif campaign == "bovin_aphto":
        ws = wb["bovin aphto"]
        ws.cell(row_idx, 5).value = int(rec["total_bovins"])
        ws.cell(row_idx, 6).value = int(rec["bovins_vaccines"])
        ws.cell(row_idx, 7).value = rec["recu_num"]
        ws.cell(row_idx, 8).value = rec["date"]
        ws.cell(row_idx, 8).number_format = "DD/MM/YYYY"
        ws.cell(row_idx, 9).value = rec["region"]
        ws.cell(row_idx, 10).value = rec["cin"]
        ws.cell(row_idx, 11).value = rec["nom"]
        
    else:  # rage
        ws = wb["داء الكلب"]
        ws.cell(row_idx, 4).value = int(rec["total_chiens"])
        ws.cell(row_idx, 5).value = int(rec["chiens_vaccines"])
        ws.cell(row_idx, 6).value = rec["recu_num"]
        ws.cell(row_idx, 7).value = rec["date"]
        ws.cell(row_idx, 7).number_format = "DD/MM/YYYY"
        ws.cell(row_idx, 8).value = rec["region"]
        ws.cell(row_idx, 9).value = rec["cin"]
        ws.cell(row_idx, 10).value = rec["nom"]
    
    wb.save(path)
    wb.close()

# ---------------------------
# NOUVELLE FONCTION: Supprimer un enregistrement
# ---------------------------
def delete_record_from_excel(path: str, campaign: str, row_idx: int):
    """Supprime un enregistrement en effaçant la ligne"""
    wb = openpyxl.load_workbook(path)
    
    if campaign == "aphto_ovin_caprin":
        ws = wb["aphto ovin et caprin"]
    elif campaign == "ovin_clavelee":
        ws = wb["ovin clavelee"]
    elif campaign == "bovin_aphto":
        ws = wb["bovin aphto"]
    else:
        ws = wb["داء الكلب"]
    
    ws.delete_rows(row_idx, 1)
    wb.save(path)
    wb.close()

        
def append_record_to_excel(path: str, campaign: str, rec: dict):
    wb = openpyxl.load_workbook(path)

    if campaign == "aphto_ovin_caprin":
        ws = wb["aphto ovin et caprin"]
        start_row = 3
        seq_col = 13
        max_col = 14

        last = find_last_data_row(ws, key_col=seq_col, start_row=start_row)
        new_row = last + 1
        copy_row_style(ws, src_row=last, dst_row=new_row, max_col=max_col)

        ws.cell(new_row, 4).value  = int(rec["total_caprins"])
        ws.cell(new_row, 5).value  = int(rec["total_ovins"])
        ws.cell(new_row, 6).value  = int(rec["caprins_vaccines"])
        ws.cell(new_row, 7).value  = int(rec["ovins_vaccines"])
        ws.cell(new_row, 8).value  = rec["recu_num"]
        ws.cell(new_row, 9).value = rec["date"]
        ws.cell(new_row, 9).number_format = "DD/MM/YYYY"
        ws.cell(new_row, 10).value = rec["region"]
        ws.cell(new_row, 11).value = rec["cin"]
        ws.cell(new_row, 12).value = rec["nom"]
        ws.cell(new_row, 13).value = int(ws.cell(last, 13).value or 0) + 1

    elif campaign == "ovin_clavelee":
        ws = wb["ovin clavelee"]
        start_row = 4
        seq_col = 10
        max_col = 10

        last = find_last_data_row(ws, key_col=seq_col, start_row=start_row)
        new_row = last + 1
        copy_row_style(ws, src_row=last, dst_row=new_row, max_col=max_col)

        ws.cell(new_row, 3).value = int(rec["total_ovins"])
        ws.cell(new_row, 4).value = int(rec["ovins_vaccines"])
        ws.cell(new_row, 5).value = rec["recu_num"]
        ws.cell(new_row, 6).value = rec["date"]
        ws.cell(new_row, 6).number_format = "DD/MM/YYYY"
        ws.cell(new_row, 7).value = rec["region"]
        ws.cell(new_row, 8).value = rec["cin"]
        ws.cell(new_row, 9).value = rec["nom"]
        ws.cell(new_row, 10).value = int(ws.cell(last, 10).value or 0) + 1

    elif campaign == "bovin_aphto":
        ws = wb["bovin aphto"]
        start_row = 4
        seq_col = 12
        max_col = 12

        last = find_last_data_row(ws, key_col=seq_col, start_row=start_row)
        new_row = last + 1
        copy_row_style(ws, src_row=last, dst_row=new_row, max_col=max_col)

        ws.cell(new_row, 5).value = int(rec["total_bovins"])
        ws.cell(new_row, 6).value = int(rec["bovins_vaccines"])
        ws.cell(new_row, 7).value = rec["recu_num"]
        ws.cell(new_row, 8).value = rec["date"]
        ws.cell(new_row, 8).number_format = "DD/MM/YYYY"
        ws.cell(new_row, 9).value = rec["region"]
        ws.cell(new_row, 10).value = rec["cin"]
        ws.cell(new_row, 11).value = rec["nom"]
        ws.cell(new_row, 12).value = int(ws.cell(last, 12).value or 0) + 1

    else:  # rage
        ws = wb["داء الكلب"]
        start_row = 5
        seq_col = 11
        max_col = 11

        last = find_last_data_row(ws, key_col=seq_col, start_row=start_row)
        new_row = last + 1
        copy_row_style(ws, src_row=last, dst_row=new_row, max_col=max_col)

        ws.cell(new_row, 4).value = int(rec["total_chiens"])
        ws.cell(new_row, 5).value = int(rec["chiens_vaccines"])
        ws.cell(new_row, 6).value = rec["recu_num"]
        ws.cell(new_row, 7).value = rec["date"]
        ws.cell(new_row, 7).number_format = "DD/MM/YYYY"
        ws.cell(new_row, 8).value = rec["region"]
        ws.cell(new_row, 9).value = rec["cin"]
        ws.cell(new_row, 10).value = rec["nom"]
        ws.cell(new_row, 11).value = int(ws.cell(last, 11).value or 0) + 1

    wb.save(path)
    wb.close()


# Configuration des options de campagne
type_options = {
    "aphto_ovin_caprin": {"label": "🐑 Fièvre Aphteuse (Ovins/Caprins)", "icon": "🐑🐐"},
    "ovin_clavelee": {"label": "🐏 Clavelée des Ovins", "icon": "🐏"},
    "bovin_aphto": {"label": "🐄 Fièvre Aphteuse (Bovins)", "icon": "🐄"},
    "rage": {"label": "🐕 Rage Canine", "icon": "🐕"},
}

# ==============================================
# INTERFACE AVEC TABS
# ==============================================

st.markdown("""
<div class="professional-form-container">
  <div class="form-header-pro">
    <div class="form-header-icon">💉</div>
    <div class="form-header-text">
      <h1 class="form-main-title" style='color :white;'>Enregistrement des Vaccinations</h1>
      <p class="form-main-subtitle">Interface professionnelle de saisie vétérinaire</p>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

# TABS
tab1, tab2 = st.tabs(["➕ Nouvelle Saisie", "✏️ Modifier/Supprimer"])

# ==============================================
# TAB 1: NOUVELLE SAISIE (code existant)
# ==============================================
with tab1:
    campaign = st.selectbox(
        "Type de campagne",
        options=list(type_options.keys()),
        format_func=lambda k: type_options[k]["label"],
        key="saisie_campaign",
    )

    campaign_label = type_options[campaign]["label"]
    campaign_icon = type_options[campaign]["icon"]

    st.info(f"Campagne active : {campaign_label}")

    st.markdown("""
        <script>
        document.addEventListener('keydown', function(e) {
        if (e.key === 'Enter' && e.target.tagName !== 'TEXTAREA') {
            e.preventDefault();
            return false;
        }
        }, true);
        </script>
        """, unsafe_allow_html=True)

    with st.form("form_saisie", clear_on_submit=True):
        
        st.markdown("""
        <div class="form-section-block">
            <div class="section-header">
                <div class="section-icon">👤</div>
                <div class="section-title">Informations Vétérinaire</div>
                <div class="section-line"></div>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        col1, col2 = st.columns(2)
        with col1:
            nom = st.text_input("👤 Nom & Prénom *", placeholder="Ex: Dr. Ahmed Bennani", key="nom_input")
            cin = st.text_input("🪪 CIN *", placeholder="Ex: AB123456", key="cin_input")
        with col2:
            region = st.text_input("📍 Région / العمادة *", placeholder="Ex: Kénitra", key="region_input")
            recu_num = st.text_input("🧾 N° Reçu *", placeholder="Ex: REC-2024-001", key="recu_input")

        st.markdown("""
        <div class="form-section-block" style="margin-top: 2rem;">
            <div class="section-header">
                <div class="section-icon">📅</div>
                <div class="section-title">Date et Données de Vaccination</div>
                <div class="section-line"></div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown('<div class="date-picker-label">📅 Date de vaccination</div>', unsafe_allow_html=True)
        date = st.date_input("Date", value=datetime.today(), label_visibility="collapsed", key="date_input")
        date_dt = pd.to_datetime(date).date()

        st.markdown("<div style='height: 1.25rem;'></div>", unsafe_allow_html=True)

        if campaign == "aphto_ovin_caprin":
            st.markdown("""
            <div class="animal-group-header">
                <span class="animal-group-icon">🐑</span>
                <span class="animal-group-title">Données Ovins</span>
            </div>
            """, unsafe_allow_html=True)

            subcol1, subcol2 = st.columns(2)
            with subcol1:
                total_ovins = st.number_input("Total ovins", min_value=0, step=1, value=0, key="total_ovins")
            with subcol2:
                ovins_vaccines = st.number_input("Ovins vaccinés", min_value=0, step=1, value=0, key="ovins_vacc")

            st.markdown("<div style='height: 1rem;'></div>", unsafe_allow_html=True)

            st.markdown("""
            <div class="animal-group-header">
                <span class="animal-group-icon">🐐</span>
                <span class="animal-group-title">Données Caprins</span>
            </div>
            """, unsafe_allow_html=True)

            subcol3, subcol4 = st.columns(2)
            with subcol3:
                total_caprins = st.number_input("Total caprins", min_value=0, step=1, value=0, key="total_caprins")
            with subcol4:
                caprins_vaccines = st.number_input("Caprins vaccinés", min_value=0, step=1, value=0, key="caprins_vacc")

        elif campaign == "ovin_clavelee":
            st.markdown("""
            <div class="animal-group-header">
                <span class="animal-group-icon">🐏</span>
                <span class="animal-group-title">Données Ovins</span>
            </div>
            """, unsafe_allow_html=True)

            subcol1, subcol2 = st.columns(2)
            with subcol1:
                total_ovins = st.number_input("Total ovins", min_value=0, step=1, value=0, key="total_ovins_clav")
            with subcol2:
                ovins_vaccines = st.number_input("Ovins vaccinés", min_value=0, step=1, value=0, key="ovins_vacc_clav")

        elif campaign == "bovin_aphto":
            st.markdown("""
            <div class="animal-group-header">
                <span class="animal-group-icon">🐄</span>
                <span class="animal-group-title">Données Bovins</span>
            </div>
            """, unsafe_allow_html=True)

            subcol1, subcol2 = st.columns(2)
            with subcol1:
                total_bovins = st.number_input("Total bovins", min_value=0, step=1, value=0, key="total_bovins")
            with subcol2:
                bovins_vaccines = st.number_input("Bovins vaccinés", min_value=0, step=1, value=0, key="bovins_vacc")

        else:  # rage
            st.markdown("""
            <div class="animal-group-header">
                <span class="animal-group-icon">🐕</span>
                <span class="animal-group-title">Données Canins</span>
            </div>
            """, unsafe_allow_html=True)

            subcol1, subcol2 = st.columns(2)
            with subcol1:
                total_chiens = st.number_input("Total chiens", min_value=0, step=1, value=0, key="total_chiens")
            with subcol2:
                chiens_vaccines = st.number_input("Chiens vaccinés", min_value=0, step=1, value=0, key="chiens_vacc")

        st.markdown('<div style="height: 2rem;"></div>', unsafe_allow_html=True)
        
        col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
        with col_btn2:
            st.markdown('<div class="form-submit-wrap">', unsafe_allow_html=True)
            submitted = st.form_submit_button("✅ Enregistrer la Vaccination", use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)

    if submitted:
        if not nom or not cin or not region or not recu_num:
            st.markdown("""
    <div class="message-box error-message">
    <div class="message-icon">⚠️</div>
    <div class="message-content">
    <div class="message-title">Champs obligatoires manquants</div>
    <div class="message-text">Veuillez remplir tous les champs marqués d'un astérisque (*)</div>
    </div>
    </div>
            """, unsafe_allow_html=True)
        else:
            rec = {
                "nom": nom,
                "cin": cin,
                "region": region,
                "recu_num": recu_num,
                "date": date_dt,
            }

            if campaign == "aphto_ovin_caprin":
                rec.update({
                    "ovins_vaccines": int(ovins_vaccines),
                    "caprins_vaccines": int(caprins_vaccines),
                    "total_ovins": int(total_ovins),
                    "total_caprins": int(total_caprins),
                })
            elif campaign == "ovin_clavelee":
                rec.update({
                    "ovins_vaccines": int(ovins_vaccines),
                    "total_ovins": int(total_ovins),
                })
            elif campaign == "bovin_aphto":
                rec.update({
                    "bovins_vaccines": int(bovins_vaccines),
                    "total_bovins": int(total_bovins),
                })
            else:
                rec.update({
                    "chiens_vaccines": int(chiens_vaccines),
                    "total_chiens": int(total_chiens),
                })

            st.markdown(f"""
    <div class="message-box success-message">
    <div class="message-icon">✅</div>
    <div class="message-content">
        <div class="message-title">Enregistrement réussi !</div>
        <div class="message-text">
            Les données de vaccination pour <strong>{nom}</strong> ont été enregistrées.
        </div>
    </div>
    </div>
            """, unsafe_allow_html=True)
            
            append_record_to_excel(DATA_FILE, campaign, rec)
            st.session_state["save_ok"] = True
            st.session_state["save_msg"] = f"✅ Données enregistrées : {nom}"
            st.cache_data.clear()
            st.rerun()

# ==============================================
# TAB 2: MODIFIER/SUPPRIMER
# ==============================================
with tab2:
    # Sélection de la campagne
    campaign_edit = st.selectbox(
        "Type de campagne",
        options=list(type_options.keys()),
        format_func=lambda k: type_options[k]["label"],
        key="edit_campaign",
    )
    
    # Charger les enregistrements
    records = load_records_from_excel(DATA_FILE, campaign_edit)
    
    if not records:
        st.warning("⚠️ Aucun enregistrement trouvé pour cette campagne.")
    else:
        st.success(f"✅ {len(records)} enregistrement(s) trouvé(s)")
        
        # Créer un DataFrame pour l'affichage
        df_records = pd.DataFrame(records)
        
        st.markdown("""
        <div class="form-section-block">
            <div class="section-header">
                <div class="section-icon">🔍</div>
                <div class="section-title">Rechercher et Modifier un Enregistrement</div>
                <div class="section-line"></div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        col_f1, col_f2, col_f3 = st.columns(3)
        
        with col_f1:
            search_nom = st.text_input("Nom vétérinaire", key="search_nom")
        with col_f2:
            search_cin = st.text_input("CIN", key="search_cin")
        with col_f3:
            search_region = st.text_input("Région", key="search_region")
        
        # Appliquer les filtres
        filtered_df = df_records.copy()
        if search_nom:
            filtered_df = filtered_df[filtered_df['nom'].str.contains(search_nom, case=False, na=False)]
        if search_cin:
            filtered_df = filtered_df[filtered_df['cin'].str.contains(search_cin, case=False, na=False)]
        if search_region:
            filtered_df = filtered_df[filtered_df['region'].str.contains(search_region, case=False, na=False)]
        
        st.markdown(f"**{len(filtered_df)} résultat(s) après filtrage**")
        
        # Afficher le tableau
        if not filtered_df.empty:
            # Sélection d'un enregistrement
            selected_seq = st.selectbox(
                "Sélectionner un enregistrement à modifier/supprimer",
                options=filtered_df['seq'].tolist(),
                format_func=lambda x: f"#{x} - {filtered_df[filtered_df['seq']==x]['nom'].iloc[0]} - {filtered_df[filtered_df['seq']==x]['date'].iloc[0]}",
                key="selected_record"
            )
            
            if selected_seq:
                selected_record = filtered_df[filtered_df['seq'] == selected_seq].iloc[0].to_dict()
                
                st.markdown("---")
                st.markdown("### ✏️ Modifier les données")
                
                # Formulaire de modification
                with st.form("form_edit"):
                    st.markdown("""
                    <div class="form-section-block">
                        <div class="section-header">
                            <div class="section-icon">👤</div>
                            <div class="section-title">Informations Vétérinaire</div>
                            <div class="section-line"></div>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        nom_edit = st.text_input("👤 Nom & Prénom *", value=selected_record["nom"], key="nom_edit")
                        cin_edit = st.text_input("🪪 CIN *", value=selected_record["cin"], key="cin_edit")
                    with col2:
                        region_edit = st.text_input("📍 Région *", value=selected_record["region"], key="region_edit")
                        recu_edit = st.text_input("🧾 N° Reçu *", value=selected_record["recu_num"], key="recu_edit")
                    
                    st.markdown("""
                    <div class="form-section-block" style="margin-top: 2rem;">
                        <div class="section-header">
                            <div class="section-icon">📅</div>
                            <div class="section-title">Date et Données</div>
                            <div class="section-line"></div>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    date_edit = st.date_input("Date", value=pd.to_datetime(selected_record["date"]), key="date_edit")
                    
                    # Champs spécifiques selon la campagne
                    if campaign_edit == "aphto_ovin_caprin":
                        st.markdown("#### 🐑 Ovins")
                        col1, col2 = st.columns(2)
                        with col1:
                            total_ovins_edit = st.number_input("Total ovins", value=int(selected_record["total_ovins"]), min_value=0, key="total_ovins_edit")
                        with col2:
                            ovins_vacc_edit = st.number_input("Ovins vaccinés", value=int(selected_record["ovins_vaccines"]), min_value=0, key="ovins_vacc_edit")
                        
                        st.markdown("#### 🐐 Caprins")
                        col3, col4 = st.columns(2)
                        with col3:
                            total_caprins_edit = st.number_input("Total caprins", value=int(selected_record["total_caprins"]), min_value=0, key="total_caprins_edit")
                        with col4:
                            caprins_vacc_edit = st.number_input("Caprins vaccinés", value=int(selected_record["caprins_vaccines"]), min_value=0, key="caprins_vacc_edit")
                    
                    elif campaign_edit == "ovin_clavelee":
                        col1, col2 = st.columns(2)
                        with col1:
                            total_ovins_edit = st.number_input("Total ovins", value=int(selected_record["total_ovins"]), min_value=0, key="total_ovins_clav_edit")
                        with col2:
                            ovins_vacc_edit = st.number_input("Ovins vaccinés", value=int(selected_record["ovins_vaccines"]), min_value=0, key="ovins_vacc_clav_edit")
                    
                    elif campaign_edit == "bovin_aphto":
                        col1, col2 = st.columns(2)
                        with col1:
                            total_bovins_edit = st.number_input("Total bovins", value=int(selected_record["total_bovins"]), min_value=0, key="total_bovins_edit")
                        with col2:
                            bovins_vacc_edit = st.number_input("Bovins vaccinés", value=int(selected_record["bovins_vaccines"]), min_value=0, key="bovins_vacc_edit")
                    
                    else:  # rage
                        col1, col2 = st.columns(2)
                        with col1:
                            total_chiens_edit = st.number_input("Total chiens", value=int(selected_record["total_chiens"]), min_value=0, key="total_chiens_edit")
                        with col2:
                            chiens_vacc_edit = st.number_input("Chiens vaccinés", value=int(selected_record["chiens_vaccines"]), min_value=0, key="chiens_vacc_edit")
                    
                    st.markdown('<div style="height: 2rem;"></div>', unsafe_allow_html=True)
                    
                    # Boutons d'action
                    col_btn1, col_btn2 = st.columns(2)
                    with col_btn1:
                        update_btn = st.form_submit_button("✅ Enregistrer les modifications", use_container_width=True)
                    with col_btn2:
                        delete_btn = st.form_submit_button("🗑️ Supprimer cet enregistrement", use_container_width=True, type="secondary")
                
                # Traitement de la modification
                if update_btn:
                    rec_update = {
                        "nom": nom_edit,
                        "cin": cin_edit,
                        "region": region_edit,
                        "recu_num": recu_edit,
                        "date": pd.to_datetime(date_edit).date(),
                    }
                    
                    if campaign_edit == "aphto_ovin_caprin":
                        rec_update.update({
                            "total_ovins": total_ovins_edit,
                            "ovins_vaccines": ovins_vacc_edit,
                            "total_caprins": total_caprins_edit,
                            "caprins_vaccines": caprins_vacc_edit,
                        })
                    elif campaign_edit == "ovin_clavelee":
                        rec_update.update({
                            "total_ovins": total_ovins_edit,
                            "ovins_vaccines": ovins_vacc_edit,
                        })
                    elif campaign_edit == "bovin_aphto":
                        rec_update.update({
                            "total_bovins": total_bovins_edit,
                            "bovins_vaccines": bovins_vacc_edit,
                        })
                    else:
                        rec_update.update({
                            "total_chiens": total_chiens_edit,
                            "chiens_vaccines": chiens_vacc_edit,
                        })
                    
                    update_record_in_excel(DATA_FILE, campaign_edit, selected_record["row_idx"], rec_update)
                    st.success(f"✅ Enregistrement #{selected_seq} modifié avec succès!")
                    st.cache_data.clear()
                    st.rerun()
                
                # Traitement de la suppression
                if delete_btn:
                    delete_record_from_excel(DATA_FILE, campaign_edit, selected_record["row_idx"])
                    st.success(f"🗑️ Enregistrement #{selected_seq} supprimé avec succès!")
                    st.cache_data.clear()
                    st.rerun()
        else:
            st.info("ℹ️ Aucun résultat ne correspond aux critères de recherche.")

# =========================
# EXPORT EXCEL (fichier complet)
# =========================

st.markdown("""
    <div class="form-section-block">
        <div class="section-header">
            <div class="section-icon">📥</div>
            <div class="section-title">Exporter le fichier Excel</div>
            <div class="section-line"></div>
        </div>
    </div>
    """, unsafe_allow_html=True)
st.markdown('<div class="download-wrap">', unsafe_allow_html=True)

with open(DATA_FILE, "rb") as f:
    st.download_button(
        label="⬇️ Télécharger mandat_sanitaire_2026.xlsx",
        data=f,
        file_name="mandat_sanitaire_2026.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key="dl_excel_full"
    )

st.markdown("</div>", unsafe_allow_html=True)

